"""
匯入 115年進項-慶豐(蔡寶琴).xlsx 到資料庫
  費用類 sheets   → company_expenses      (公司費用)
  會計師類 sheets → company_input_invoices (工程進項)

費用 sheet 三種格式：
  A (113.3-4 etc.)  : 日期,發票號碼,開立發票人,廠商統一編號,品名,銷售額,稅額,其他,總計,備註
  B (113.11-12 etc.): 日期,發票號碼,開立發票人,品名,銷售額,稅額,其他,總計,備註
  C (115.1-2 etc.)  : 日期,發票號碼(字母),發票號碼(數字),開立發票人,品名,銷售額,稅額,其他,總計,備註

會計師 sheet 兩種格式：
  OLD (113.x-x): 材料別,日期,發票字號,廠商名稱,材料名稱,金額,稅金,總金額,備註,地點
  NEW (114.9-10+): 材料別,日期,發票字號(字母),統編,廠商名稱,材料名稱,金額,稅金,總金額,備註,地點,[分攤1],[分攤2]

重複判斷：
  一般發票: (invoice_no, date, total_amount) 三者相同 → 跳過
  收據:     (date, vendor_name, total_amount) 三者相同 → 跳過

用法: python scripts/import_cindy_data.py [--dry-run]
"""
import sys, io, re, datetime
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
import psycopg2
from psycopg2.extras import execute_values

DRY_RUN   = '--dry-run' in sys.argv
XLSX_PATH = 'd:/erp_cindy/ocr-service/115年進項-慶豐(蔡寶琴).xlsx'
DB_URL    = 'postgresql://postgres:EqnGlCVFOtVhPVAvdlPcojEGKCTbVmlF@interchange.proxy.rlwy.net:33523/railway'

LOCATION_MAP = {
    '東明段108': 1,
    '東明段8':   4,
    '平和段607': 7,
}

def detect_project(loc):
    if not loc: return None
    s = str(loc)
    for key, pid in LOCATION_MAP.items():
        if key in s:
            return pid
    return None

def to_f(v):
    try:
        s = str(v).strip() if v is not None else ''
        return float(s) if s not in ('', 'None', ' ') else 0.0
    except (ValueError, TypeError):
        return 0.0

def parse_roc_date(val):
    if val is None: return None
    if isinstance(val, datetime.datetime): return val.strftime('%Y-%m-%d')
    if isinstance(val, datetime.date):     return val.strftime('%Y-%m-%d')
    s = re.sub(r'[/年月]', '', str(val).strip())
    if re.match(r'^\d{7}$', s):
        try: return datetime.date(int(s[:3])+1911, int(s[3:5]), int(s[5:7])).strftime('%Y-%m-%d')
        except ValueError: return None
    if re.match(r'^\d{6}$', s):
        try: return datetime.date(int(s[:3])+1911, int(s[3:5]), int(s[5:7])).strftime('%Y-%m-%d')
        except ValueError: return None
    # try 5-digit without leading zero in month: 113819 → handled above, but catch '11389'
    if re.match(r'^\d{5}$', s):
        try: return datetime.date(int(s[:3])+1911, int(s[3:4]), int(s[4:5])).strftime('%Y-%m-%d')
        except ValueError: return None
    return None

def sheet_period(sname):
    sname = sname.strip()
    m = re.search(r'(\d{3})[.\-年](\d{1,2})[-月](\d{1,2})', sname)
    if m:
        return f"{m.group(1)}.{int(m.group(2))}-{int(m.group(3))}"
    m = re.match(r'(\d{3}[.\-]\d+[-]\d+)', sname)
    if m: return m.group(1)
    return sname[:10].strip()

def str_or_none(v):
    if v is None: return None
    s = str(v).strip()
    return s if s not in ('', 'None') else None

# ─────────────────────────────────────────────
# 費用 sheet 解析（三種格式）
# ─────────────────────────────────────────────
def detect_expense_fmt(ws):
    """A=舊版(含統編欄), B=中版(無統編), C=新版(None分隔符)"""
    for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
        if row[0] is None: continue
        if '日期' not in str(row[0]): continue
        col2 = str(row[2]).strip() if row[2] is not None else None
        col3 = str(row[3]).strip() if row[3] is not None else None
        if col2 is None:
            return 'C'
        if col3 and ('編號' in col3 or '統一' in col3 or '廠商' in col3):
            return 'A'
        return 'B'
    return 'B'

def parse_expense_sheet(ws, period):
    fmt = detect_expense_fmt(ws)
    rows = []

    # 找 header row 序號
    hdr_idx = 1
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True), start=1):
        if row[0] and '日期' in str(row[0]):
            hdr_idx = i
            break

    skip_kw = {'日期', '收據', '電子發票', '三聯式', '二聯式', '合計', '小計',
                '慶豐', '', 'None', '日期（民國）', '113年', '114年', '115年'}

    for row in ws.iter_rows(min_row=hdr_idx+1, values_only=True):
        if row[0] is None: continue
        v0 = str(row[0]).strip()
        if v0 in skip_kw: continue
        # 跳過年度扣繳憑單行
        if '年' in v0 and '/' not in v0 and len(v0) <= 5: continue

        exp_date = parse_roc_date(row[0])
        if exp_date is None: continue

        if fmt == 'A':
            invoice_no  = str_or_none(row[1])
            vendor_name = str_or_none(row[2])
            vendor_tid  = str_or_none(row[3])
            item_name   = str_or_none(row[4])
            amt         = to_f(row[5])
            tax         = to_f(row[6])
            other       = to_f(row[7])
            total       = to_f(row[8])
            note        = str_or_none(row[9]) if len(row) > 9 else None
        elif fmt == 'B':
            invoice_no  = str_or_none(row[1])
            vendor_name = str_or_none(row[2])
            vendor_tid  = None
            item_name   = str_or_none(row[3])
            amt         = to_f(row[4])
            tax         = to_f(row[5])
            other       = to_f(row[6])
            total       = to_f(row[7])
            note        = str_or_none(row[8]) if len(row) > 8 else None
        else:  # C
            alpha       = str_or_none(row[1]) or ''
            num         = str_or_none(row[2]) or ''
            invoice_no  = (alpha + num).strip() or None
            vendor_name = str_or_none(row[3])
            vendor_tid  = None
            item_name   = str_or_none(row[4])
            amt         = to_f(row[5])
            tax         = to_f(row[6])
            other       = to_f(row[7])
            total       = to_f(row[8])
            note        = str_or_none(row[9]) if len(row) > 9 else None

        if total == 0: continue
        if '要注意' in (note or ''): note = None

        rows.append({
            'expense_date': exp_date,
            'invoice_no':   invoice_no,
            'invoice_type': None,
            'vendor_tax_id': None,
            'vendor_name':  vendor_name,
            'item_name':    item_name,
            'amount':       amt,
            'tax_amount':   tax,
            'other_amount': other,
            'total_amount': total,
            'period':       period,
            'note':         note,
        })
    return rows

# ─────────────────────────────────────────────
# 會計師 sheet 解析（兩種格式）
# ─────────────────────────────────────────────
def detect_acc_fmt(ws):
    """OLD=col3廠商名稱 / NEW=col3 None(統編)"""
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        if row[0] is None: continue
        if '材料別' not in str(row[0]): continue
        col3 = row[3]
        return 'NEW' if col3 is None else 'OLD'
    return 'OLD'

def parse_accountant_sheet(ws, period):
    fmt = detect_acc_fmt(ws)
    rows = []

    hdr_idx = 1
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        if row[0] and '材料別' in str(row[0]):
            hdr_idx = i
            break

    skip_kw = {'材料別', '合計', '小計', '', 'None'}

    for row in ws.iter_rows(min_row=hdr_idx+1, values_only=True):
        if row[0] is None: continue
        mat = str(row[0]).strip()
        if mat in skip_kw: continue

        date_val = row[1]
        inv_date = parse_roc_date(date_val)
        if inv_date is None: continue

        if fmt == 'OLD':
            # col2=全發票號碼, col3=廠商名稱, col4=材料名稱
            # col5=金額, col6=稅金, col7=總金額, col8=備註, col9=地點
            invoice_no  = str_or_none(row[2])
            vendor_tid  = None
            vendor_name = str_or_none(row[3])
            item_name   = str_or_none(row[4])
            amt         = to_f(row[5])
            tax         = to_f(row[6])
            total       = to_f(row[7])
            note_v      = str_or_none(row[8])
            location    = str_or_none(row[9]) if len(row) > 9 else None
            split1 = split2 = None
        else:
            # col2=字母, col3=統編, col4=廠商名稱, col5=材料名稱
            # col6=金額, col7=稅金, col8=總金額, col9=備註, col10=地點, col11,12=分攤
            alpha       = str_or_none(row[2]) or ''
            tid         = str_or_none(row[3]) or ''
            invoice_no  = f"{alpha}-{tid}".strip('-') or None
            vendor_tid  = tid or None
            vendor_name = str_or_none(row[4])
            item_name   = str_or_none(row[5])
            amt         = to_f(row[6])
            tax         = to_f(row[7])
            total       = to_f(row[8])
            note_v      = str_or_none(row[9])
            location    = str_or_none(row[10]) if len(row) > 10 else None
            split1      = to_f(row[11]) if len(row) > 11 and row[11] is not None else None
            split2      = to_f(row[12]) if len(row) > 12 and row[12] is not None else None

        if total == 0: continue

        # 跨兩地分攤
        if split1 and split2 and location and ('及' in location or '和' in location):
            parts = re.split(r'及|和', location)
            loc1 = parts[0].strip()
            loc2 = parts[1].strip() if len(parts) > 1 else location
            pid1 = detect_project(loc1)
            pid2 = detect_project(loc2)
            ratio1 = split1 / total if total else 0
            tax1 = round(tax * ratio1)
            tax2 = round(tax - tax1)
            rows.append({'invoice_date':inv_date,'invoice_no':invoice_no,'vendor_tax_id':vendor_tid,
                         'vendor_name':vendor_name,'material_type':mat,'item_name':item_name,
                         'amount':round(split1-tax1),'tax_amount':tax1,'total_amount':split1,
                         'project_id':pid1,'location':loc1,'period':period,'note':note_v})
            rows.append({'invoice_date':inv_date,'invoice_no':invoice_no,'vendor_tax_id':vendor_tid,
                         'vendor_name':vendor_name,'material_type':mat,'item_name':item_name,
                         'amount':round(split2-tax2),'tax_amount':tax2,'total_amount':split2,
                         'project_id':pid2,'location':loc2,'period':period,'note':note_v})
        else:
            pid = detect_project(location)
            rows.append({'invoice_date':inv_date,'invoice_no':invoice_no,'vendor_tax_id':vendor_tid,
                         'vendor_name':vendor_name,'material_type':mat,'item_name':item_name,
                         'amount':amt,'tax_amount':tax,'total_amount':total,
                         'project_id':pid,'location':location,'period':period,'note':note_v})
    return rows

# ─────────────────────────────────────────────
def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    all_expenses = []
    all_invoices = []

    for sname in wb.sheetnames:
        sc = sname.strip()
        period = sheet_period(sc)
        ws = wb[sname]

        if '費用' in sc:
            rows = parse_expense_sheet(ws, period)
            print(f'[費用] {sc}: {len(rows)} 筆')
            all_expenses.extend(rows)
        elif '會計師' in sc:
            rows = parse_accountant_sheet(ws, period)
            print(f'[會計] {sc}: {len(rows)} 筆')
            all_invoices.extend(rows)

    print(f'\n費用原始總筆數: {len(all_expenses)}')
    print(f'進項原始總筆數: {len(all_invoices)}')

    # ── 去重 費用 ──
    seen_exp = set()
    dedup_exp = []
    dup_exp = 0
    for e in all_expenses:
        inv = e['invoice_no'] or ''
        is_receipt = not inv or inv.startswith('收據') or re.match(r'^[A-Za-z\-]*$', inv)
        if is_receipt:
            key = ('R', e['expense_date'] or '', str(e['vendor_name'] or ''), str(e['total_amount']))
        else:
            key = ('I', inv, e['expense_date'] or '', str(e['total_amount']))
        if key in seen_exp:
            dup_exp += 1
            print(f'  [重複費用] {key}')
        else:
            seen_exp.add(key)
            dedup_exp.append(e)

    # ── 去重 進項 ──
    seen_inv = set()
    dedup_inv = []
    dup_inv = 0
    for i in all_invoices:
        inv = i['invoice_no'] or ''
        key = (inv, i['invoice_date'] or '', str(i['total_amount']), str(i['project_id'] or ''))
        if key in seen_inv:
            dup_inv += 1
            print(f'  [重複進項] {key}')
        else:
            seen_inv.add(key)
            dedup_inv.append(i)

    print(f'\n費用去重後: {len(dedup_exp)} 筆 (跳過 {dup_exp} 筆)')
    print(f'進項去重後: {len(dedup_inv)} 筆 (跳過 {dup_inv} 筆)')

    if DRY_RUN:
        print('\n[DRY-RUN] 未寫入資料庫')
        print('\n費用前10筆:')
        for e in dedup_exp[:10]:
            print(f"  {e['expense_date']}  {str(e['vendor_name'] or '')[:20]:20s}  {e['total_amount']:>10}  {e['period']}")
        print('\n進項前10筆:')
        for i in dedup_inv[:10]:
            print(f"  {i['invoice_date']}  {str(i['vendor_name'] or '')[:20]:20s}  {i['total_amount']:>10}  pid={i['project_id']}  {i['period']}")
        return

    # ── 寫入 DB ──
    conn = psycopg2.connect(DB_URL)
    cur  = conn.cursor()

    now = datetime.datetime.now()

    if dedup_exp:
        exp_cols = ['expense_date','invoice_no','invoice_type','vendor_tax_id','vendor_name',
                    'item_name','amount','tax_amount','other_amount','total_amount','period','note',
                    'created_at','updated_at']
        execute_values(cur,
            f"INSERT INTO company_expenses ({','.join(exp_cols)}) VALUES %s",
            [tuple(list(e.get(c) for c in exp_cols[:-2]) + [now, now]) for e in dedup_exp])
        print(f'已寫入 company_expenses: {len(dedup_exp)} 筆')

    if dedup_inv:
        inv_cols = ['invoice_date','invoice_no','vendor_tax_id','vendor_name','material_type',
                    'item_name','amount','tax_amount','total_amount','project_id','location','period','note',
                    'created_at','updated_at']
        execute_values(cur,
            f"INSERT INTO company_input_invoices ({','.join(inv_cols)}) VALUES %s",
            [tuple(list(i.get(c) for c in inv_cols[:-2]) + [now, now]) for i in dedup_inv])
        print(f'已寫入 company_input_invoices: {len(dedup_inv)} 筆')

    conn.commit()
    cur.close()
    conn.close()
    print('\n匯入完成！')

if __name__ == '__main__':
    main()
